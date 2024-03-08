# Project Team 3 (Alexander Smith, Daniel Smith, John Sharp)
# Ivy Tech Asset Manager - File Manager
# Provides functions to handle operations around the files containing asset inventories.

import openpyxl
import os
from campus import Campus
from building import Building
from equipment import Equipment

class FileManager:
    # Load a file into memory
    def load_file(self, file_path):
        file_name = os.path.basename(file_path)
        campus = Campus(location=file_name.split(".")[0])
        current_file = openpyxl.load_workbook(file_path)

        for sheet in current_file.get_sheet_names():
            # Get the sheet
            current_sheet = current_file[sheet]
            # Make a building instance from the sheet
            building = Building(building_name=sheet)
            # Add building's items to the building
            for row in range(1, current_sheet.max_row):
                item_info = []
                for col in current_sheet.iter_cols(0, current_sheet.max_column):
                    item_info.append(col[row].value)
                building.add_item(Equipment(item_name=item_info[0], room_quantity={item_info[1] : int(item_info[3])}, 
                                            department=item_info[2], manufacturer=item_info[4], description=item_info[5]))
            campus.add_building(building)
        return campus

    # Create a new blank site file with the given buildings
    def create_file(self, file_path, buildings):
        # Create a workbook document
        wb = openpyxl.Workbook()
        # Fill in the initial worksheet
        ws = wb.active
        ws.title = buildings[0]
        # Set header fields
        ws["A1"] = "Name"
        ws["B1"] = "Location"
        ws["C1"] = "Department"
        ws["D1"] = "Quantity"
        ws["E1"] = "Manufacturer"
        ws["F1"] = "Description"

        # We can loop through creating the remaining worksheets, if any
        for i in range(1, len(buildings)):
            new_sheet = wb.create_sheet(title=buildings[i])
            # Set header fields
            new_sheet["A1"] = "Name"
            new_sheet["B1"] = "Location"
            new_sheet["C1"] = "Department"
            new_sheet["D1"] = "Quantity"
            new_sheet["E1"] = "Manufacturer"
            new_sheet["F1"] = "Description"
            
        # Now we can save the file
        wb.save(file_path)

        return


    # TODO save_file(self, file_path, ...)
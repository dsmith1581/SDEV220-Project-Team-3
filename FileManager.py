import openpyxl
import os
from Campus import Campus
from Building import Building
from Equipment import Equipment

class FileManager:
    
    def loadFile(self, file_path):
        file_name = os.path.basename(file_path)
        campusInfo = file_name.split("-")
        campus = Campus(location=campusInfo[0], id=campusInfo[1])
        current_file = openpyxl.load_workbook(file_path)
        for sheet in current_file.get_sheet_names():
            # Get the sheet
            current_sheet = current_file[sheet]
            # Make a building instance from the sheet
            building = Building(building_name=sheet)
            # Add building's items to the building
            for row in range(1, current_sheet.max_row):
                item_info = []
                for col in current_sheet.iter_cols(0, current_sheet.max_column):
                    item_info.append(col[row].value)
                building.add_item(Equipment(item_name=item_info[0], room_quantity={item_info[1] : int(item_info[3])}, 
                                            department=item_info[2], manufacturer=item_info[4], description=item_info[5]))
            campus.add_building(building)
        return campus